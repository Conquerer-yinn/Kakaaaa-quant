"""存储层：多 Sheet 重写、表格区域与总览刷新。"""
import os

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

import storage.excel_helper as excel_helper
from storage.excel_helper import ExcelHelper


@pytest.fixture
def base_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(excel_helper, "MASTER_DATA_DIR", str(tmp_path / "master"))
    monkeypatch.setattr(excel_helper, "BACKUP_DIR", str(tmp_path / "backups"))
    return str(tmp_path / "master")


def _sample_df():
    return pd.DataFrame({"日期": ["20260101", "20260102"], "涨停数": [50, 60]})


def test_upsert_creates_sheets_and_tables(base_dir):
    path = ExcelHelper.upsert_data_workbook(
        "wb.xlsx",
        sheets={"总市场数据": _sample_df(), "高度观察": _sample_df()},
        table_names={"总市场数据": "tbl_market"},
        base_dir=base_dir,
    )

    workbook = load_workbook(path)
    assert {"总市场数据", "高度观察"} <= set(workbook.sheetnames)
    assert "tbl_market" in workbook["总市场数据"].tables
    assert workbook["总市场数据"]["A1"].value == "日期"


def test_upsert_preserves_template_sheet(base_dir):
    os.makedirs(base_dir, exist_ok=True)
    template = Workbook()
    sheet = template.active
    sheet.title = "图表模板"
    sheet["A1"] = "keep-me"
    template.save(os.path.join(base_dir, "wb.xlsx"))

    ExcelHelper.upsert_data_workbook("wb.xlsx", sheets={"总市场数据": _sample_df()}, base_dir=base_dir)

    workbook = load_workbook(os.path.join(base_dir, "wb.xlsx"))
    assert "图表模板" in workbook.sheetnames
    assert workbook["图表模板"]["A1"].value == "keep-me"


def test_update_overview_sheet(base_dir):
    rows = [["近期市场情绪总览"], ["最新日期", "20260102", None, "运行模式", "test"]]
    path = ExcelHelper.update_overview_sheet("wb.xlsx", "总览", rows, base_dir=base_dir)

    workbook = load_workbook(path)
    sheet = workbook["总览"]
    assert sheet["A1"].value == "近期市场情绪总览"
    assert sheet["A1"].font.bold is True
    assert sheet["B2"].value == "20260102"
