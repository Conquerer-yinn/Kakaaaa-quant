import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from strategies.chinext_limit_up_event_study import EventStudyResult
from strategies.chinext_limit_up_workbook import (
    build_event_study_file_name,
    find_latest_event_study_workbook,
    write_event_study_workbook,
)


def sample_result() -> EventStudyResult:
    return EventStudyResult(
        details=pd.DataFrame(
            [
                {
                    "事件日期": "20260105",
                    "股票代码": "300001.SZ",
                    "股票名称": "样本一",
                    "5日收益率(%)": 10.0,
                }
            ]
        ),
        summary=pd.DataFrame(
            [{"观察周期": "5日", "样本数": 1, "平均收益率(%)": 10.0}]
        ),
        candidate_event_count=2,
        complete_sample_count=1,
        skipped_incomplete_count=1,
        skipped_missing_quote_count=0,
        missing_benchmark_count=0,
        excluded_st_count=0,
        excluded_recent_listing_count=0,
        missing_stock_basic_count=0,
        group_summary=pd.DataFrame(
            [
                {
                    "分组维度": "连板阶段",
                    "分组": "首板",
                    "观察周期": "5日",
                    "样本数": 1,
                    "平均收益率(%)": 10.0,
                    "平均超额收益率(%)": 5.0,
                }
            ]
        ),
        quality_summary=pd.DataFrame(
            [
                {"质量项目": "候选事件", "数量": 2, "说明": "创业板且涨停"},
                {"质量项目": "完整样本", "数量": 1, "说明": "进入统计"},
            ]
        ),
    )


class ChinextLimitUpWorkbookTest(unittest.TestCase):
    def test_builds_stable_file_name(self):
        self.assertEqual(
            build_event_study_file_name("20260101", "20260331"),
            "创业板涨停事件研究_20260101_20260331.xlsx",
        )

    def test_writes_five_required_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = write_event_study_workbook(
                sample_result(),
                start_date="20260101",
                end_date="20260331",
                base_dir=temp_dir,
            )

            workbook = load_workbook(output_path, read_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["研究摘要", "分组统计", "样本质量", "事件明细", "运行信息"],
                )
            finally:
                workbook.close()
            run_info = pd.read_excel(output_path, sheet_name="运行信息")
            info = dict(zip(run_info["字段"], run_info["值"]))
            self.assertEqual(info["研究开始日期"], "20260101")
            self.assertEqual(info["研究结束日期"], "20260331")
            self.assertEqual(info["候选事件数"], 2)
            self.assertEqual(info["完整样本数"], 1)

    def test_finds_latest_workbook_by_end_date(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            older = Path(temp_dir) / build_event_study_file_name("20250101", "20251231")
            newer = Path(temp_dir) / build_event_study_file_name("20260101", "20260331")
            ignored = Path(temp_dir) / "其他文件.xlsx"
            older.touch()
            newer.touch()
            ignored.touch()

            latest = find_latest_event_study_workbook(base_dir=temp_dir)

            self.assertIsNotNone(latest)
            self.assertEqual(latest.file_name, newer.name)
            self.assertEqual(latest.start_date, "20260101")
            self.assertEqual(latest.end_date, "20260331")
            self.assertEqual(latest.file_path, str(newer))

    def test_returns_none_when_result_directory_has_no_matching_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            self.assertIsNone(find_latest_event_study_workbook(base_dir=temp_dir))


if __name__ == "__main__":
    unittest.main()
