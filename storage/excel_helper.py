import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from common.config import BACKUP_DIR, MASTER_DATA_DIR


class ExcelHelper:
    """Excel 落表辅助函数。"""

    @staticmethod
    def ensure_storage_dirs():
        os.makedirs(MASTER_DATA_DIR, exist_ok=True)
        os.makedirs(BACKUP_DIR, exist_ok=True)

    @staticmethod
    def build_storage_path(file_name, base_dir):
        ExcelHelper.ensure_storage_dirs()
        os.makedirs(base_dir, exist_ok=True)
        return os.path.join(base_dir, file_name)

    @staticmethod
    def build_master_path(file_name):
        return ExcelHelper.build_storage_path(file_name, MASTER_DATA_DIR)

    @staticmethod
    def build_backup_path(file_name):
        return ExcelHelper.build_storage_path(file_name, BACKUP_DIR)

    @staticmethod
    def backup_file(file_path):
        if not os.path.exists(file_path):
            return None

        ExcelHelper.ensure_storage_dirs()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{timestamp}_{os.path.basename(file_path)}"
        backup_path = os.path.join(BACKUP_DIR, backup_name)
        shutil.copy2(file_path, backup_path)
        return backup_path

    @staticmethod
    def read_sheet(file_name, sheet_name, base_dir=MASTER_DATA_DIR):
        file_path = ExcelHelper.build_storage_path(file_name, base_dir)
        if not os.path.exists(file_path):
            return None

        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except ValueError:
            # 旧工作簿如果还没有这个 sheet，就按首次写入处理。
            return None
        return ExcelHelper._normalize_date_column(df)

    @staticmethod
    def append_rows(df, file_name, sheet_name="Sheet1", dedupe_subset=None, base_dir=MASTER_DATA_DIR):
        if df.empty:
            raise ValueError("Cannot write an empty DataFrame to Excel.")

        file_path = ExcelHelper.build_storage_path(file_name, base_dir)
        if os.path.exists(file_path):
            ExcelHelper.backup_file(file_path)
            existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
            merged_df = pd.concat([existing_df, df], ignore_index=True)
        else:
            merged_df = df.copy()

        if dedupe_subset:
            merged_df = merged_df.drop_duplicates(subset=dedupe_subset, keep="last")

        merged_df = ExcelHelper._normalize_date_column(merged_df)

        try:
            merged_df.to_excel(file_path, index=False, sheet_name=sheet_name)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot write to {file_path}. The file may be open in Excel. Close it and run the job again."
            ) from exc

        return file_path

    @staticmethod
    def upsert_data_workbook(file_name, sheets, table_names=None, base_dir=MASTER_DATA_DIR):
        """只更新指定数据 sheet，保留同一本工作簿里的模板 sheet / 图表 sheet。"""
        if not sheets:
            raise ValueError("Cannot create a workbook without sheets.")

        file_path = ExcelHelper.build_storage_path(file_name, base_dir)
        if os.path.exists(file_path):
            ExcelHelper.backup_file(file_path)
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)

        for sheet_name, df in sheets.items():
            worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
            normalized_df = ExcelHelper._normalize_date_column(df)
            ExcelHelper._reset_worksheet(worksheet)
            ExcelHelper._write_dataframe(worksheet, normalized_df)

            if table_names and sheet_name in table_names:
                ExcelHelper._ensure_table(worksheet, table_names[sheet_name])

        for worksheet in workbook.worksheets:
            ExcelHelper._autofit_columns(worksheet)

        try:
            workbook.save(file_path)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot write to {file_path}. The file may be open in Excel. Close it and run the job again."
            ) from exc

        return file_path

    @staticmethod
    def _normalize_date_column(df):
        """统一把日期列保留成 YYYYMMDD 字符串，避免 Excel/引擎再次写坏日期。"""
        if df is None or df.empty or "日期" not in df.columns:
            return df

        export_df = df.copy()
        export_df["日期"] = export_df["日期"].map(ExcelHelper._normalize_single_date)
        return export_df

    @staticmethod
    def _normalize_single_date(value):
        if pd.isna(value):
            return value
        if isinstance(value, datetime):
            return value.strftime("%Y%m%d")

        text = str(value).strip()
        if not text:
            return text
        if text.isdigit() and len(text) == 8:
            return text
        if text.endswith('.0') and text[:-2].isdigit() and len(text[:-2]) == 8:
            return text[:-2]

        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%Y%m%d")
        return text

